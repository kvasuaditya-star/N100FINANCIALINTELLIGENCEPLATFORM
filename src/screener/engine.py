import logging
import os
import sqlite3

import numpy as np
import pandas as pd
import yaml

logging.basicConfig(
    level=logging.INFO, format="%(asctime)s - %(levelname)s - %(message)s"
)
logger = logging.getLogger(__name__)


def load_config(config_path="config/screener_config.yaml"):
    with open(config_path, "r") as f:
        return yaml.safe_load(f)


def compute_composite_score(df, df_cf):
    """
    Computes a sector-relative composite quality score (0-100).
    Normalizes metrics within each broad_sector using P10/P90 winsorization.
    """
    df = df.copy()

    # 1. Calculate missing derived metrics
    # CFO / PAT
    df["cfo_pat_ratio"] = np.where(
        df["net_profit"] > 0, df["cash_from_operations_cr"] / df["net_profit"], 0
    )

    # FCF Positive Flag
    df["fcf_positive_flag"] = np.where(df["free_cash_flow_cr"] > 0, 1, 0)

    # FCF CAGR (5yr)
    fcf_cagrs = []
    for _, row in df.iterrows():
        cid = row["company_id"]
        year = row["year"]
        try:
            year_num = int(year[:4])
            start_year = str(year_num - 5)

            end_cf = df_cf[
                (df_cf["company_id"] == cid)
                & (df_cf["year"].str.startswith(str(year_num)))
            ]
            start_cf = df_cf[
                (df_cf["company_id"] == cid)
                & (df_cf["year"].str.startswith(start_year))
            ]

            if not end_cf.empty and not start_cf.empty:
                end_fcf = (
                    end_cf["operating_activity"].values[0]
                    + end_cf["investing_activity"].values[0]
                )
                start_fcf = (
                    start_cf["operating_activity"].values[0]
                    + start_cf["investing_activity"].values[0]
                )

                if start_fcf > 0 and end_fcf > 0:
                    cagr = ((end_fcf / start_fcf) ** (1.0 / 5.0) - 1.0) * 100
                else:
                    cagr = 0
            else:
                cagr = 0
        except Exception:
            cagr = 0
        fcf_cagrs.append(cagr)

    df["fcf_cagr_5yr"] = fcf_cagrs

    # Pre-fill NaNs with 0 or group medians for scoring purposes
    metrics = [
        "return_on_equity_pct",
        "roce",
        "net_profit_margin_pct",
        "fcf_cagr_5yr",
        "cfo_pat_ratio",
        "fcf_positive_flag",
        "revenue_cagr_5yr",
        "pat_cagr_5yr",
        "debt_to_equity",
        "interest_coverage",
    ]
    for m in metrics:
        if m in df.columns:
            df[m] = df[m].fillna(df[m].median())

    # Treat infinite ICR (Debt Free) as a large number
    df["interest_coverage"] = np.where(
        df["icr_label"] == "Debt Free", 100.0, df["interest_coverage"]
    )

    df["composite_quality_score"] = 0.0

    # Normalise within each sector
    for sector in df["broad_sector"].unique():
        mask = df["broad_sector"] == sector
        if mask.sum() == 0:
            continue

        sector_df = df[mask].copy()
        scores = {}

        for m in metrics:
            series = sector_df[m].astype(float)

            # P10/P90 Winsorisation
            p10 = series.quantile(0.10)
            p90 = series.quantile(0.90)

            if pd.isna(p10) or pd.isna(p90):
                p10, p90 = series.min(), series.max()

            winsorized = series.clip(p10, p90)

            # Min-Max Scaling (0 to 100)
            min_val = winsorized.min()
            max_val = winsorized.max()

            if max_val > min_val:
                scaled = (winsorized - min_val) / (max_val - min_val) * 100.0
            else:
                scaled = pd.Series(50.0, index=winsorized.index)

            # Invert D/E (lower is better)
            if m == "debt_to_equity":
                scaled = 100.0 - scaled

            scores[m] = scaled

        # Apply Weights
        # Profitability (35%): ROE 15%, ROCE 10%, NPM 10%
        prof_score = (
            scores["return_on_equity_pct"] * 0.15
            + scores["roce"] * 0.10
            + scores["net_profit_margin_pct"] * 0.10
        )

        # Cash Quality (30%): FCF CAGR 15%, CFO/PAT 10%, FCF flag 5%
        cash_score = (
            scores["fcf_cagr_5yr"] * 0.15
            + scores["cfo_pat_ratio"] * 0.10
            + scores["fcf_positive_flag"] * 0.05
        )

        # Growth (20%): Rev CAGR 10%, PAT CAGR 10%
        growth_score = scores["revenue_cagr_5yr"] * 0.10 + scores["pat_cagr_5yr"] * 0.10

        # Leverage (15%): D/E 10%, ICR 5%
        lev_score = scores["debt_to_equity"] * 0.10 + scores["interest_coverage"] * 0.05

        total_score = prof_score + cash_score + growth_score + lev_score

        # Scale final score so that max is 100 within sector
        if total_score.max() > total_score.min():
            total_score = (
                (total_score - total_score.min())
                / (total_score.max() - total_score.min())
                * 100.0
            )

        df.loc[mask, "composite_quality_score"] = total_score

    return df


def apply_filters(df, df_fr_hist, filters):
    """
    Applies the specified filters to the dataframe.
    """
    filtered = df.copy()

    for key, value in filters.items():
        if key == "roe_min":
            filtered = filtered[filtered["return_on_equity_pct"] >= value]
        elif key == "de_max":
            # Skip companies in Financials sector for D/E max filter
            mask = (filtered["debt_to_equity"] <= value) | (
                filtered["broad_sector"] == "Financials"
            )
            filtered = filtered[mask]
        elif key == "fcf_min":
            filtered = filtered[filtered["free_cash_flow_cr"] >= value]
        elif key == "revenue_cagr_5yr_min":
            filtered = filtered[filtered["revenue_cagr_5yr"] >= value]
        elif key == "pat_cagr_5yr_min":
            filtered = filtered[filtered["pat_cagr_5yr"] >= value]
        elif key == "pe_max":
            filtered = filtered[filtered["pe_ratio"] <= value]
        elif key == "pb_max":
            filtered = filtered[filtered["pb_ratio"] <= value]
        elif key == "dividend_yield_min":
            filtered = filtered[filtered["dividend_yield_pct"] >= value]
        elif key == "dividend_payout_max":
            filtered = filtered[filtered["dividend_payout_ratio_pct"] <= value]
        elif key == "sales_min":
            filtered = filtered[filtered["sales"] >= value]
        elif key == "de_declining" and value is True:
            # Check if D/E is declining YoY
            declining_mask = []
            for _, row in filtered.iterrows():
                cid = row["company_id"]
                year = row["year"]
                current_de = row["debt_to_equity"]
                try:
                    year_num = int(year[:4])
                    prev_year = str(year_num - 1)
                    prev_de = df_fr_hist[
                        (df_fr_hist["company_id"] == cid)
                        & (df_fr_hist["year"].str.startswith(prev_year))
                    ]["debt_to_equity"]
                    if not prev_de.empty:
                        declining_mask.append(current_de < prev_de.values[0])
                    else:
                        declining_mask.append(False)
                except Exception:
                    declining_mask.append(False)
            filtered = filtered[declining_mask]
        elif key == "revenue_cagr_3yr_min":
            filtered = filtered[filtered["revenue_cagr_3yr"] >= value]

    return filtered


def generate_screener_output(db_path="data/nifty100.db"):
    if not os.path.exists(db_path):
        logger.error(f"Database not found at {db_path}")
        return

    conn = sqlite3.connect(db_path)

    query = """
    WITH LatestYear AS (
        SELECT company_id, MAX(year) as latest_year
        FROM financial_ratios
        GROUP BY company_id
    )
    SELECT 
        fr.*,
        c.company_name, c.roce_percentage as roce, 
        s.broad_sector, s.sub_sector,
        mc.pe_ratio, mc.pb_ratio, mc.dividend_yield_pct, mc.market_cap_crore,
        pl.sales, pl.net_profit,
        cf.operating_activity, cf.investing_activity
    FROM financial_ratios fr
    JOIN LatestYear ly ON fr.company_id = ly.company_id AND fr.year = ly.latest_year
    JOIN companies c ON fr.company_id = c.id
    LEFT JOIN sectors s ON fr.company_id = s.company_id
    LEFT JOIN market_cap mc ON fr.company_id = mc.company_id AND CAST(SUBSTR(fr.year, 1, 4) AS INTEGER) = mc.year
    LEFT JOIN profitandloss pl ON fr.company_id = pl.company_id AND fr.year = pl.year
    LEFT JOIN cashflow cf ON fr.company_id = cf.company_id AND fr.year = cf.year
    """

    df = pd.read_sql_query(query, conn)
    df_cf = pd.read_sql_query(
        "SELECT company_id, year, operating_activity, investing_activity FROM cashflow",
        conn,
    )
    df_fr_hist = pd.read_sql_query(
        "SELECT company_id, year, debt_to_equity FROM financial_ratios", conn
    )

    conn.close()

    df = compute_composite_score(df, df_cf)

    config = load_config()
    presets = config.get("presets", {})

    os.makedirs("output", exist_ok=True)
    output_file = "output/screener_output.xlsx"

    kpi_columns = [
        "company_id",
        "company_name",
        "broad_sector",
        "composite_quality_score",
        "return_on_equity_pct",
        "roce",
        "debt_to_equity",
        "free_cash_flow_cr",
        "revenue_cagr_5yr",
        "pat_cagr_5yr",
        "net_profit_margin_pct",
        "pe_ratio",
        "pb_ratio",
        "dividend_yield_pct",
        "dividend_payout_ratio_pct",
        "interest_coverage",
        "asset_turnover",
        "sales",
        "market_cap_crore",
        "revenue_cagr_3yr",
    ]

    with pd.ExcelWriter(output_file, engine="openpyxl") as writer:
        for preset_id, preset_data in presets.items():
            name = preset_data.get("name", preset_id)
            filters = preset_data.get("filters", {})

            filtered_df = apply_filters(df, df_fr_hist, filters)
            filtered_df = filtered_df.sort_values(
                by="composite_quality_score", ascending=False
            )

            # Select KPI columns
            available_cols = [c for c in kpi_columns if c in filtered_df.columns]
            export_df = filtered_df[available_cols].copy()

            # Format composite score
            export_df["composite_quality_score"] = export_df[
                "composite_quality_score"
            ].round(2)

            # Write to excel sheet
            # Max sheet name length is 31
            sheet_name = name[:31]
            export_df.to_excel(writer, sheet_name=sheet_name, index=False)

            logger.info(f"Preset '{name}': {len(export_df)} companies found.")

    # Apply conditional formatting using openpyxl
    import openpyxl
    from openpyxl.styles import PatternFill

    wb = openpyxl.load_workbook(output_file)
    green_fill = PatternFill(
        start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"
    )
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

    for preset_id, preset_data in presets.items():
        name = preset_data.get("name", preset_id)
        sheet_name = name[:31]
        ws = wb[sheet_name]

        filters = preset_data.get("filters", {})

        # Map filter keys to column indices
        col_map = {
            col[0].value: col[0].column for col in ws.iter_cols(min_row=1, max_row=1)
        }

        filter_col_mapping = {
            "roe_min": "return_on_equity_pct",
            "de_max": "debt_to_equity",
            "fcf_min": "free_cash_flow_cr",
            "revenue_cagr_5yr_min": "revenue_cagr_5yr",
            "pat_cagr_5yr_min": "pat_cagr_5yr",
            "pe_max": "pe_ratio",
            "pb_max": "pb_ratio",
            "dividend_yield_min": "dividend_yield_pct",
            "dividend_payout_max": "dividend_payout_ratio_pct",
            "sales_min": "sales",
            "revenue_cagr_3yr_min": "revenue_cagr_3yr",
        }

        for row_idx in range(2, ws.max_row + 1):
            for filter_key, threshold in filters.items():
                col_name = filter_col_mapping.get(filter_key)
                if not col_name or col_name not in col_map:
                    continue

                col_idx = col_map[col_name]
                cell = ws.cell(row=row_idx, column=col_idx)

                if cell.value is None:
                    continue

                try:
                    val = float(cell.value)
                    passed = False
                    if filter_key.endswith("_min"):
                        passed = val >= threshold
                    elif filter_key.endswith("_max"):
                        # Exception for financials on D/E
                        if (
                            filter_key == "de_max"
                            and ws.cell(
                                row=row_idx, column=col_map["broad_sector"]
                            ).value
                            == "Financials"
                        ):
                            passed = True
                        else:
                            passed = val <= threshold

                    if passed:
                        cell.fill = green_fill
                    else:
                        cell.fill = red_fill
                except (ValueError, TypeError):
                    pass

    wb.save(output_file)
    logger.info(f"Screener output saved to {output_file} with conditional formatting.")


if __name__ == "__main__":
    generate_screener_output()
