import logging
import os
import sqlite3

import matplotlib.pyplot as plt
import numpy as np
import pandas as pd

logging.basicConfig(
    level=logging.INFO, format="%(asctime)s - %(levelname)s - %(message)s"
)
logger = logging.getLogger(__name__)


def compute_peer_percentiles(db_path="data/nifty100.db"):
    if not os.path.exists(db_path):
        logger.error(f"Database not found at {db_path}")
        return None, None

    conn = sqlite3.connect(db_path)

    # We want latest year data from financial_ratios
    query = """
    WITH LatestYear AS (
        SELECT company_id, MAX(year) as latest_year
        FROM financial_ratios
        GROUP BY company_id
    )
    SELECT 
        fr.*,
        c.company_name, c.roce_percentage as roce, 
        s.broad_sector, s.sub_sector,
        mc.pe_ratio, mc.pb_ratio, mc.dividend_yield_pct, mc.market_cap_crore,
        pl.sales, pl.net_profit,
        cf.operating_activity, cf.investing_activity,
        pg.peer_group_name, pg.is_benchmark
    FROM financial_ratios fr
    JOIN LatestYear ly ON fr.company_id = ly.company_id AND fr.year = ly.latest_year
    JOIN companies c ON fr.company_id = c.id
    LEFT JOIN sectors s ON fr.company_id = s.company_id
    LEFT JOIN market_cap mc ON fr.company_id = mc.company_id AND CAST(SUBSTR(fr.year, 1, 4) AS INTEGER) = mc.year
    LEFT JOIN profitandloss pl ON fr.company_id = pl.company_id AND fr.year = pl.year
    LEFT JOIN cashflow cf ON fr.company_id = cf.company_id AND fr.year = cf.year
    LEFT JOIN peer_groups pg ON fr.company_id = pg.company_id
    """

    df = pd.read_sql_query(query, conn)

    # 10 metrics to rank
    metrics_to_rank = {
        "return_on_equity_pct": "ROE",
        "roce": "ROCE",
        "net_profit_margin_pct": "Net Profit Margin",
        "debt_to_equity": "D/E",
        "free_cash_flow_cr": "FCF",
        "pat_cagr_5yr": "PAT CAGR 5yr",
        "revenue_cagr_5yr": "Revenue CAGR 5yr",
        "eps_cagr_5yr": "EPS CAGR 5yr",
        "interest_coverage": "Interest Coverage",
        "asset_turnover": "Asset Turnover",
    }

    # Treat 'Debt Free' ICR as 100.0 for ranking
    df["interest_coverage"] = np.where(
        df["icr_label"] == "Debt Free", 100.0, df["interest_coverage"]
    )

    # Create table for percentiles
    conn.execute("""
    CREATE TABLE IF NOT EXISTS peer_percentiles (
        company_id VARCHAR,
        peer_group_name VARCHAR,
        metric VARCHAR,
        value NUMERIC,
        percentile_rank NUMERIC,
        year VARCHAR,
        PRIMARY KEY (company_id, metric, year),
        FOREIGN KEY (company_id) REFERENCES companies(id)
    )
    """)
    conn.execute("DELETE FROM peer_percentiles")
    conn.commit()

    percentile_records = []

    for _, row in df.iterrows():
        cid = row["company_id"]
        pg = row["peer_group_name"]
        year = row["year"]

        if pd.isna(pg):
            # No peer group assigned
            for col, metric_name in metrics_to_rank.items():
                percentile_records.append(
                    {
                        "company_id": cid,
                        "peer_group_name": "No peer group assigned",
                        "metric": metric_name,
                        "value": row[col] if not pd.isna(row[col]) else None,
                        "percentile_rank": None,
                        "year": year,
                    }
                )
            continue

        # Has peer group - compute percentiles
        # We need the peer group dataframe for the current metric
        pg_df = df[df["peer_group_name"] == pg]

        for col, metric_name in metrics_to_rank.items():
            val = row[col]
            if pd.isna(val):
                pct_rank = None
            else:
                # Calculate percentile rank within peer group (0 to 1)
                series = pg_df[col].dropna()
                if len(series) > 1:
                    pct_rank = (series < val).sum() / (len(series) - 1)
                else:
                    pct_rank = 1.0  # Only one company or all others NA

                # Invert for D/E
                if col == "debt_to_equity" and pct_rank is not None:
                    pct_rank = 1.0 - pct_rank

            percentile_records.append(
                {
                    "company_id": cid,
                    "peer_group_name": pg,
                    "metric": metric_name,
                    "value": val,
                    "percentile_rank": pct_rank,
                    "year": year,
                }
            )

    df_percentiles = pd.DataFrame(percentile_records)

    # Insert to database
    df_percentiles.to_sql("peer_percentiles", conn, if_exists="append", index=False)
    conn.commit()
    conn.close()

    logger.info(f"Populated peer_percentiles table with {len(df_percentiles)} records.")
    return df, df_percentiles


def generate_peer_comparison_report(df, df_percentiles):
    os.makedirs("output", exist_ok=True)
    output_file = "output/peer_comparison.xlsx"

    peer_groups = df["peer_group_name"].dropna().unique()

    # We want to export sheet per peer group
    metrics = {
        "return_on_equity_pct": "ROE",
        "roce": "ROCE",
        "net_profit_margin_pct": "Net Profit Margin",
        "debt_to_equity": "D/E",
        "free_cash_flow_cr": "FCF",
        "pat_cagr_5yr": "PAT CAGR 5yr",
        "revenue_cagr_5yr": "Revenue CAGR 5yr",
        "eps_cagr_5yr": "EPS CAGR 5yr",
        "interest_coverage": "Interest Coverage",
        "asset_turnover": "Asset Turnover",
    }

    with pd.ExcelWriter(output_file, engine="openpyxl") as writer:
        for pg in peer_groups:
            if pg == "No peer group assigned":
                continue

            pg_df = df[df["peer_group_name"] == pg].copy()

            # Prepare output dataframe
            cols = ["company_id", "company_name", "is_benchmark"]
            out_df = pg_df[cols].copy()

            # Add metric values and ranks
            for col, metric_name in metrics.items():
                out_df[metric_name] = pg_df[col]

                # Fetch ranks
                ranks = []
                for cid in out_df["company_id"]:
                    rank_val = df_percentiles[
                        (df_percentiles["company_id"] == cid)
                        & (df_percentiles["metric"] == metric_name)
                    ]["percentile_rank"].values
                    if len(rank_val) > 0 and not pd.isna(rank_val[0]):
                        ranks.append(rank_val[0] * 100.0)
                    else:
                        ranks.append(None)
                out_df[f"{metric_name} Rank"] = ranks

            # Add Summary Row (Median)
            summary_row = {
                "company_id": "SUMMARY",
                "company_name": "Peer Group Median",
                "is_benchmark": 0,
            }
            for col, metric_name in metrics.items():
                summary_row[metric_name] = pg_df[col].median()
                summary_row[f"{metric_name} Rank"] = None

            out_df = pd.concat([out_df, pd.DataFrame([summary_row])], ignore_index=True)

            # Write sheet
            sheet_name = pg[:31]
            out_df.to_excel(writer, sheet_name=sheet_name, index=False)

    # Apply styling
    import openpyxl
    from openpyxl.styles import PatternFill

    wb = openpyxl.load_workbook(output_file)

    green_fill = PatternFill(
        start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"
    )
    yellow_fill = PatternFill(
        start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"
    )
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    gold_fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")

    for pg in peer_groups:
        if pg == "No peer group assigned":
            continue

        sheet_name = pg[:31]
        ws = wb[sheet_name]

        # Highlight benchmark row
        benchmark_col_idx = None
        for c_idx, cell in enumerate(ws[1], 1):
            if cell.value == "is_benchmark":
                benchmark_col_idx = c_idx
                break

        for row in range(2, ws.max_row + 1):
            if (
                benchmark_col_idx
                and ws.cell(row=row, column=benchmark_col_idx).value == 1
            ):
                for col in range(1, ws.max_column + 1):
                    ws.cell(row=row, column=col).fill = gold_fill

            # Check percentile ranks for color coding
            for col in range(1, ws.max_column + 1):
                header = ws.cell(row=1, column=col).value
                if header and header.endswith(" Rank"):
                    val = ws.cell(row=row, column=col).value
                    if val is not None:
                        try:
                            val = float(val)
                            # Avoid overwriting gold fill on benchmark row if we want it completely gold,
                            # but the requirement says "Colour-code percentile rank cells". So we will overwrite.
                            if val >= 75:
                                ws.cell(row=row, column=col).fill = green_fill
                            elif val > 25:
                                ws.cell(row=row, column=col).fill = yellow_fill
                            else:
                                ws.cell(row=row, column=col).fill = red_fill
                        except ValueError:
                            pass

        # Hide the is_benchmark column
        if benchmark_col_idx:
            col_letter = openpyxl.utils.get_column_letter(benchmark_col_idx)
            ws.column_dimensions[col_letter].hidden = True

    wb.save(output_file)
    logger.info(f"Peer comparison output saved to {output_file}")


def generate_radar_charts(df, df_percentiles):
    os.makedirs("reports/radar_charts", exist_ok=True)

    metrics = {
        "return_on_equity_pct": "ROE",
        "roce": "ROCE",
        "net_profit_margin_pct": "NPM",
        "debt_to_equity": "D/E",
        # We need FCF Score but the requirement says FCF Score. We'll use FCF value or rank?
        # Requirement: "8 axes: ROE, ROCE, NPM, D/E, FCF score, PAT CAGR 5yr, Revenue CAGR 5yr, Composite Score"
        # We don't have a single "FCF score" in df, we have "free_cash_flow_cr". We'll use FCF.
        "free_cash_flow_cr": "FCF",
        "pat_cagr_5yr": "PAT CAGR 5yr",
        "revenue_cagr_5yr": "Revenue CAGR 5yr",
        "composite_quality_score": "Composite Score",
    }

    # We need to compute composite_quality_score if not present.
    # Let's import engine and run compute_composite_score.
    import sys

    sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
    from screener.engine import compute_composite_score

    # We also need df_cf for composite score
    conn = sqlite3.connect("data/nifty100.db")
    df_cf = pd.read_sql_query(
        "SELECT company_id, year, operating_activity, investing_activity FROM cashflow",
        conn,
    )
    conn.close()

    df = compute_composite_score(df, df_cf)

    peer_groups = df["peer_group_name"].dropna().unique()

    # Nifty 100 Averages for companies without peer group
    nifty_avgs = {k: df[k].median() for k in metrics}

    for _, row in df.iterrows():
        cid = row["company_id"]
        pg = row["peer_group_name"]

        # Get company values
        comp_vals = [
            row.get(k, 0) if not pd.isna(row.get(k, 0)) else 0 for k in metrics
        ]

        # Get reference values
        if pd.isna(pg):
            ref_vals = [nifty_avgs[k] for k in metrics]
            ref_label = "Nifty 100 Median"
        else:
            pg_df = df[df["peer_group_name"] == pg]
            ref_vals = [
                pg_df[k].median() if not pd.isna(pg_df[k].median()) else 0
                for k in metrics
            ]
            ref_label = f"{pg} Median"

        # We need to normalize values for radar chart (0 to 1) so axes are comparable
        # Find max across Nifty100 for each metric to normalize
        normalized_comp = []
        normalized_ref = []
        for i, k in enumerate(metrics.keys()):
            max_val = df[k].max()
            min_val = df[k].min()
            if max_val > min_val:
                norm_c = (comp_vals[i] - min_val) / (max_val - min_val)
                norm_r = (ref_vals[i] - min_val) / (max_val - min_val)

                # Invert D/E
                if k == "debt_to_equity":
                    norm_c = 1.0 - norm_c
                    norm_r = 1.0 - norm_r
            else:
                norm_c = 0.5
                norm_r = 0.5
            normalized_comp.append(norm_c)
            normalized_ref.append(norm_r)

        # Draw radar chart
        angles = np.linspace(0, 2 * np.pi, len(metrics), endpoint=False).tolist()
        normalized_comp += normalized_comp[:1]
        normalized_ref += normalized_ref[:1]
        angles += angles[:1]

        fig, ax = plt.subplots(figsize=(6, 6), subplot_kw=dict(polar=True))
        ax.fill(angles, normalized_comp, color="blue", alpha=0.25)
        ax.plot(angles, normalized_comp, color="blue", linewidth=2, label=cid)

        ax.plot(
            angles,
            normalized_ref,
            color="orange",
            linewidth=2,
            linestyle="dashed",
            label=ref_label,
        )

        ax.set_xticks(angles[:-1])
        ax.set_xticklabels(list(metrics.values()), size=8)
        ax.set_yticklabels([])  # Hide radial ticks

        plt.legend(loc="upper right", bbox_to_anchor=(1.3, 1.1))
        plt.title(f"{row['company_name']} ({cid})", size=11, y=1.1)

        plt.savefig(f"reports/radar_charts/{cid}_radar.png", bbox_inches="tight")
        plt.close()

    logger.info("Generated radar charts for all companies.")


if __name__ == "__main__":
    df, df_percentiles = compute_peer_percentiles()
    if df is not None:
        generate_peer_comparison_report(df, df_percentiles)
        generate_radar_charts(df, df_percentiles)
