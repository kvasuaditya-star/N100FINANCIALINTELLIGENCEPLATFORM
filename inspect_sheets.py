import os
import openpyxl

raw_dir = "data/raw"
for file in sorted(os.listdir(raw_dir)):
    if file.endswith(".xlsx"):
        path = os.path.join(raw_dir, file)
        try:
            wb = openpyxl.load_workbook(path, read_only=True)
            print(f"{file}: Sheets = {wb.sheetnames}")
            # Try to see some content
            sheet = wb[wb.sheetnames[0]]
            rows = list(sheet.iter_rows(max_row=3, values_only=True))
            print(f"  Row 0: {rows[0] if len(rows) > 0 else 'None'}")
            print(f"  Row 1: {rows[1] if len(rows) > 1 else 'None'}")
            print(f"  Row 2: {rows[2] if len(rows) > 2 else 'None'}")
        except Exception as e:
            print(f"Error reading {file}: {e}")
